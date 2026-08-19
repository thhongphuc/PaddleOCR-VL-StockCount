"""
Stock Count Multi-Format Exporter
Generates styled Excel (.xlsx), CSV, JSON, and Markdown export reports.
"""

import io
import json
import csv
from typing import Dict, List, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class StockCountExporter:
    @staticmethod
    def export_excel(data: Dict[str, Any]) -> bytes:
        """
        Generate a professional, color-coded Excel spreadsheet for Stock Count Audit.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock Count Audit"

        # Ensure grid lines are visible
        ws.views.sheetView[0].showGridLines = True

        # Styles definition
        title_font = Font(name="Calibri", size=16, bold=True, color="1E293B")
        section_font = Font(name="Calibri", size=11, bold=True, color="334155")
        meta_label_font = Font(name="Calibri", size=10, bold=True, color="475569")
        meta_value_font = Font(name="Calibri", size=10, color="0F172A")
        
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Deep Navy Blue
        
        data_font = Font(name="Calibri", size=10, color="0F172A")
        bold_data_font = Font(name="Calibri", size=10, bold=True, color="0F172A")
        
        thin_border_side = Side(border_style="thin", color="CBD5E1")
        table_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        # Status Fills
        match_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid") # Light Emerald
        match_font = Font(name="Calibri", size=10, bold=True, color="047857")
        
        surplus_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Light Amber
        surplus_font = Font(name="Calibri", size=10, bold=True, color="B45309")
        
        deficit_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Light Red
        deficit_font = Font(name="Calibri", size=10, bold=True, color="B91C1C")

        # 1. Main Title
        ws.merge_cells("A1:K1")
        title_cell = ws["A1"]
        title_cell.value = "BIÊN BẢN KIỂM KÊ TỒN KHO • STOCK COUNT AUDIT REPORT"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        metadata = data.get("metadata", {})
        kpi = data.get("kpi", {})
        items = data.get("items", [])

        # 2. Metadata Block (Rows 3-5)
        meta_fields = [
            ("Kho hàng (Warehouse):", metadata.get("warehouse", "Kho Tổng"), "Mã phiếu (Sheet No):", metadata.get("document_no", "SC-2026")),
            ("Ngày kiểm kê (Date):", metadata.get("count_date", ""), "Người kiểm (Auditor):", metadata.get("auditor", "")),
            ("Hình thức (Type):", metadata.get("count_type", "Kiểm kê định kỳ"), "Tỷ lệ khớp tồn kho:", f"{kpi.get('match_rate_pct', 100)}% (Khớp {kpi.get('matched_skus', 0)}/{kpi.get('total_skus', 0)} SKU)")
        ]

        for r_idx, row_vals in enumerate(meta_fields, start=3):
            ws[f"A{r_idx}"].value = row_vals[0]
            ws[f"A{r_idx}"].font = meta_label_font
            ws[f"B{r_idx}"].value = row_vals[1]
            ws[f"B{r_idx}"].font = meta_value_font
            ws.merge_cells(f"B{r_idx}:D{r_idx}")

            ws[f"F{r_idx}"].value = row_vals[2]
            ws[f"F{r_idx}"].font = meta_label_font
            ws[f"G{r_idx}"].value = row_vals[3]
            ws[f"G{r_idx}"].font = bold_data_font if "Tỷ lệ" in row_vals[2] else meta_value_font
            ws.merge_cells(f"G{r_idx}:J{r_idx}")
            ws.row_dimensions[r_idx].height = 20

        # 3. Table Headers (Row 7)
        headers = [
            ("STT", 6, "center"),
            ("Mã SKU / Barcode", 18, "left"),
            ("Tên Hàng Hóa / Quy Cách", 32, "left"),
            ("ĐVT", 8, "center"),
            ("Vị Trí / Kệ", 14, "center"),
            ("Tồn Sổ Sách", 14, "right"),
            ("Thực Tế Đếm", 14, "right"),
            ("Chênh Lệch", 14, "right"),
            ("Trạng Thái", 18, "center"),
            ("Mã Lô / Batch", 15, "center"),
            ("Hạn Dùng (EXP)", 14, "center"),
            ("Ghi Chú / Tình Trạng", 24, "left")
        ]

        start_row = 7
        ws.row_dimensions[start_row].height = 28
        for col_idx, (h_name, col_w, align) in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            cell = ws[f"{col_letter}{start_row}"]
            cell.value = h_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            cell.border = table_border
            ws.column_dimensions[col_letter].width = col_w

        # 4. Data Rows
        current_row = start_row + 1
        for item in items:
            ws.row_dimensions[current_row].height = 22
            
            stt = item.get("stt", current_row - start_row)
            sku = item.get("sku", "")
            desc = item.get("description", "")
            uom = item.get("uom", "")
            loc = item.get("location", "")
            book_qty = item.get("book_qty", 0)
            actual_qty = item.get("actual_qty", 0)
            variance = item.get("variance", 0)
            status = item.get("status", "MATCHED")
            status_text = item.get("status_text", "Khớp")
            lot = item.get("lot_batch", "")
            exp = item.get("expiry", "")
            remarks = item.get("remarks", "")

            row_data = [
                (stt, "center", data_font, None),
                (sku, "left", bold_data_font, None),
                (desc, "left", data_font, None),
                (uom, "center", data_font, None),
                (loc, "center", data_font, None),
                (book_qty, "right", data_font, None),
                (actual_qty, "right", bold_data_font, None),
                (variance, "right", None, None), # Will format based on status
                (status_text, "center", None, None), # Will format based on status
                (lot, "center", data_font, None),
                (exp, "center", data_font, None),
                (remarks, "left", data_font, None),
            ]

            # Choose status format
            if status == "MATCHED" or abs(variance) < 1e-4:
                status_f = match_font
                status_b = match_fill
            elif status == "SURPLUS" or variance > 0:
                status_f = surplus_font
                status_b = surplus_fill
            else:
                status_f = deficit_font
                status_b = deficit_fill

            for col_idx, (val, align, f_style, b_style) in enumerate(row_data, start=1):
                col_letter = get_column_letter(col_idx)
                cell = ws[f"{col_letter}{current_row}"]
                cell.value = val
                cell.alignment = Alignment(horizontal=align, vertical="center")
                cell.border = table_border

                if col_idx in [8, 9]: # Variance & Status columns
                    cell.font = status_f
                    cell.fill = status_b
                else:
                    cell.font = f_style
                    if b_style:
                        cell.fill = b_style

                # Number formatting for quantities
                if col_idx in [6, 7, 8]:
                    cell.number_format = "#,##0"

            current_row += 1

        # 5. Total Row
        ws.row_dimensions[current_row].height = 25
        total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        total_font = Font(name="Calibri", size=11, bold=True, color="0F172A")

        ws.merge_cells(f"A{current_row}:E{current_row}")
        total_label_cell = ws[f"A{current_row}"]
        total_label_cell.value = f"TỔNG CỘNG ({len(items)} SKU)"
        total_label_cell.font = total_font
        total_label_cell.fill = total_fill
        total_label_cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx in range(1, 13):
            col_letter = get_column_letter(col_idx)
            c = ws[f"{col_letter}{current_row}"]
            c.border = table_border
            c.fill = total_fill

        # Book Total Formula
        book_total_cell = ws[f"F{current_row}"]
        book_total_cell.value = f"=SUM(F{start_row+1}:F{current_row-1})"
        book_total_cell.font = total_font
        book_total_cell.alignment = Alignment(horizontal="right", vertical="center")
        book_total_cell.number_format = "#,##0"

        # Actual Total Formula
        actual_total_cell = ws[f"G{current_row}"]
        actual_total_cell.value = f"=SUM(G{start_row+1}:G{current_row-1})"
        actual_total_cell.font = total_font
        actual_total_cell.alignment = Alignment(horizontal="right", vertical="center")
        actual_total_cell.number_format = "#,##0"

        # Variance Total Formula
        var_total_cell = ws[f"H{current_row}"]
        var_total_cell.value = f"=SUM(H{start_row+1}:H{current_row-1})"
        var_total_cell.font = total_font
        var_total_cell.alignment = Alignment(horizontal="right", vertical="center")
        var_total_cell.number_format = "+#,##0;-#,##0;0"

        # 6. Signatures block (Rows current_row + 3)
        sig_row = current_row + 3
        signatures = [
            ("Người Lập Phiếu\n(Ký & ghi rõ họ tên)", "B"),
            ("Thủ Kho Quản Lý\n(Ký & ghi rõ họ tên)", "E"),
            ("Kiểm Toán / Giám Sát\n(Ký & ghi rõ họ tên)", "H"),
            ("Ban Giám Đốc / Kế Toán Trưởng\n(Ký & đóng dấu)", "K"),
        ]
        for title, col_let in signatures:
            cell = ws[f"{col_let}{sig_row}"]
            cell.value = title
            cell.font = bold_data_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def export_csv(data: Dict[str, Any]) -> str:
        """Export Stock Count data to standard CSV string"""
        output = io.StringIO()
        # Add UTF-8 BOM for Excel UTF-8 display compatibility
        output.write("\ufeff")
        
        writer = csv.writer(output, lineterminator="\n")
        
        # Header Info
        meta = data.get("metadata", {})
        writer.writerow(["BIÊN BẢN KIỂM KÊ TỒN KHO - STOCK COUNT AUDIT"])
        writer.writerow(["Kho hàng", meta.get("warehouse", "")])
        writer.writerow(["Mã phiếu", meta.get("document_no", "")])
        writer.writerow(["Ngày kiểm kê", meta.get("count_date", "")])
        writer.writerow(["Người kiểm", meta.get("auditor", "")])
        writer.writerow([]) # blank row

        # Table Header
        writer.writerow([
            "STT", "Mã SKU", "Tên Hàng Hóa", "ĐVT", "Vị Trí",
            "Tồn Sổ Sách", "Thực Tế Đếm", "Chênh Lệch", "Trạng Thái",
            "Mã Lô", "Hạn Dùng", "Ghi Chú"
        ])

        for item in data.get("items", []):
            writer.writerow([
                item.get("stt", ""),
                item.get("sku", ""),
                item.get("description", ""),
                item.get("uom", ""),
                item.get("location", ""),
                item.get("book_qty", 0),
                item.get("actual_qty", 0),
                item.get("variance", 0),
                item.get("status_text", ""),
                item.get("lot_batch", ""),
                item.get("expiry", ""),
                item.get("remarks", "")
            ])

        return output.getvalue()

    @staticmethod
    def export_json(data: Dict[str, Any]) -> str:
        """Export Stock Count data to JSON string"""
        return json.dumps(data, ensure_ascii=False, indent=2)

    @staticmethod
    def export_markdown(data: Dict[str, Any]) -> str:
        """Export Stock Count data to Github-flavored Markdown table"""
        meta = data.get("metadata", {})
        kpi = data.get("kpi", {})
        items = data.get("items", [])

        lines = [
            f"# 📋 BIÊN BẢN KIỂM KÊ TỒN KHO ({meta.get('document_no', 'SC-2026')})",
            "",
            f"- **Kho hàng**: {meta.get('warehouse', 'Kho Tổng')}",
            f"- **Ngày kiểm kê**: {meta.get('count_date', '')}",
            f"- **Người kiểm kê**: {meta.get('auditor', '')}",
            f"- **Tỷ lệ khớp tồn kho**: {kpi.get('match_rate_pct', 100)}% ({kpi.get('matched_skus', 0)}/{kpi.get('total_skus', 0)} SKU khớp)",
            "",
            "| STT | Mã SKU | Tên Hàng Hóa | ĐVT | Vị Trí | Tồn Sổ | Thực Tế | Chênh Lệch | Trạng Thái | Mã Lô | Hạn Dùng | Ghi Chú |",
            "| :---: | :--- | :--- | :---: | :---: | ---: | ---: | ---: | :---: | :---: | :---: | :--- |"
        ]

        for item in items:
            stt = item.get("stt", "")
            sku = item.get("sku", "")
            desc = item.get("description", "").replace("|", "-")
            uom = item.get("uom", "")
            loc = item.get("location", "")
            book = item.get("book_qty", 0)
            actual = item.get("actual_qty", 0)
            var = item.get("variance", 0)
            status = item.get("status_text", "")
            lot = item.get("lot_batch", "")
            exp = item.get("expiry", "")
            rem = item.get("remarks", "").replace("|", "-")

            lines.append(f"| {stt} | `{sku}` | {desc} | {uom} | {loc} | {book} | **{actual}** | {var:+d} | {status} | {lot} | {exp} | {rem} |")

        lines.append("")
        lines.append(f"**Tổng số lượng sổ sách:** {kpi.get('total_book_qty', 0)} | **Tổng thực tế:** {kpi.get('total_actual_qty', 0)} | **Chênh lệch ròng:** {kpi.get('net_variance_units', 0):+d}")
        return "\n".join(lines)
